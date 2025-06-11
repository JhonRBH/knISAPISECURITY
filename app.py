from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

RUTA_ARCHIVO = r"C:\Users\jhon.rangel\Downloads\bd.xlsx"

@app.route('/guardar', methods=['POST'])
def guardar():
    data = request.get_json()
    validados_shp = data.get('validados', [])
    errores = data.get('errores', [])
    datos = data.get('datos', [])

    if os.path.exists(RUTA_ARCHIVO):
        wb = load_workbook(RUTA_ARCHIVO)
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)

    # Crear hojas si no existen
    if 'Validados' not in wb.sheetnames:
        wb.create_sheet('Validados')
    if 'Errores' not in wb.sheetnames:
        wb.create_sheet('Errores')

    hoja_validados = wb['Validados']
    hoja_errores = wb['Errores']

    # Si está vacía la hoja validados, colocar encabezados
    if hoja_validados.max_row == 1:
        hoja_validados.append(['Ordnum', 'SHP', 'LOAD', 'WD', 'Carrier'])

    # Guardar registros validados: para cada SHP validado busco su fila original
    for shp in validados_shp:
        fila = next((f for f in datos if str(f.get('SHP')).strip() == shp), None)
        if fila:
            hoja_validados.append([
                fila.get('Ordnum', ''),
                fila.get('SHP', ''),
                fila.get('LOAD', ''),
                fila.get('WD', ''),
                fila.get('Carrier', ''),
            ])

    # Encabezado para errores
    if hoja_errores.max_row == 1:
        hoja_errores.append(['SHP', 'Tipo Error'])

    for error in errores:
        hoja_errores.append([
            error.get('shp', ''),
            error.get('tipo_error', '')
        ])

    wb.save(RUTA_ARCHIVO)

    return jsonify({"message": "✅ Registros guardados en bd.xlsx correctamente."})

if __name__ == '__main__':
    app.run(port=5001, debug=True)
